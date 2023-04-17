import time
from openpyxl import load_workbook
from selenium import webdriver
from selenium.webdriver.common.keys import Keys

# Load data from the Excel file
wb = load_workbook("data.xlsx")
ws = wb.active

# Set up the web driver and navigate to the form page
driver = webdriver.Chrome()
driver.get("https://example.com/form")

# Iterate through the rows in the Excel file (skipping the header row)
for row in ws.iter_rows(min_row=2, values_only=True):
    first_name, last_name, email = row

    # Fill out the form fields
    first_name_field = driver.find_element_by_name("first_name")
    first_name_field.send_keys(first_name)

    last_name_field = driver.find_element_by_name("last_name")
    last_name_field.send_keys(last_name)

    email_field = driver.find_element_by_name("email")
    email_field.send_keys(email)

    # Submit the form
    submit_button = driver.find_element_by_xpath("//input[@type='submit']")
    submit_button.click()

    # Wait for the form to process (adjust the time as needed)
    time.sleep(5)

    # Navigate back to the form page to enter the next row of data
    driver.get("https://example.com/form")

# Close the browser window
driver.quit()